from flask import Flask, request, render_template, redirect,url_for
from openpyxl import load_workbook
from openpyxl.styles import Protection
import os

app = Flask(__name__)
EXCEL_FILE = "data.xlsx"

@app.route('/')
def index():
    success = request.args.get("success") == "true"
    return render_template("form.html", success=success)


@app.route('/submit', methods=['POST'])
def submit():
    # Get data from form
    data = {
        "entry_type": request.form.get("entry_type"),
        "date": request.form.get("date"),
        "project_name": request.form.get("project_name"),
        "amount_by": request.form.get("amount_by"),
        "amount": f"₹{request.form.get('amount')}",
        "invoice_number": request.form.get("invoice_number", ""),
        "party_name": request.form.get("party_name", ""),
        "expense_for": request.form.get("expense_for", ""),
        "description": request.form.get("description", ""),
        "labour_name": request.form.get("labour_name", "")
    }

    # Load Excel
    if not os.path.exists(EXCEL_FILE):
        return "Excel file not found. Please run setup_excel.py first."

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Append data to next row
    row = [
        data["entry_type"], data["date"], data["project_name"], data["amount_by"],
        data["amount"], data["invoice_number"], data["party_name"],
        data["expense_for"], data["description"], data["labour_name"]
    ]
    ws.append(row)

    # Lock empty fields
    row_index = ws.max_row
    for col_index, value in enumerate(row, start=1):
        cell = ws.cell(row=row_index, column=col_index)
        if not value:
            cell.protection = Protection(locked=True)

    # Protect sheet
    ws.protection.sheet = True
    ws.protection.password = "1234"

    # Save
    wb.save(EXCEL_FILE)

    return redirect(url_for('index', success='true'))


if __name__ == '__main__':
    app.run(debug=True)
